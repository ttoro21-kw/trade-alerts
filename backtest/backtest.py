"""
백테스트 스크립트 — 지난 2년 데이터로 임계값 검증.

각 종목에 대해:
  1) 최근 2년 데이터 fetch
  2) 매일 trailing peak/trough 갱신하며 신호 검출
  3) 신호 발생 시 가상의 "매도/매수" 시뮬레이션
  4) 결과를 Excel로 출력 (per-trade + 종목별 summary + per-ticker chart)

사용법:
  python backtest.py                # 전 종목, 24개월
  python backtest.py --period 12mo  # 12개월
  python backtest.py --ticker NVDA  # 특정 종목만
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import yfinance as yf

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from thresholds import (  # noqa: E402
    TICKER_CONFIG,
    PEAK_TROUGH_LOOKBACK_DAYS,
    EMA_SMOOTHING_DAYS,
)


def fetch_history(ticker: str, period: str = "2y") -> pd.DataFrame:
    df = yf.download(ticker, period=period, interval="1d",
                     progress=False, auto_adjust=True)
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    return df


def compute_atr(df: pd.DataFrame, period: int = 14) -> pd.Series:
    h, l, c = df['High'], df['Low'], df['Close']
    pc = c.shift(1)
    tr = pd.concat([h - l, (h - pc).abs(), (l - pc).abs()], axis=1).max(axis=1)
    return tr.rolling(period).mean()


def simulate(df: pd.DataFrame, ticker: str) -> pd.DataFrame:
    """
    각 거래일마다:
      - 5일 EMA 종가, rolling 40일 peak/trough 계산
      - 손절/매수 단계 결정
      - 단계가 처음 도달하는 날만 'event'로 기록
      - 'open trade' 트래킹: 매수 신호 후 다음 손절 신호까지 보유 시 손익 계산
    """
    cfg = TICKER_CONFIG[ticker]
    ts, tb = cfg["trailing_stop"], cfg["trailing_buy"]

    close_ema = df['Close'].ewm(span=EMA_SMOOTHING_DAYS, adjust=False).mean()
    rolling_peak = close_ema.rolling(PEAK_TROUGH_LOOKBACK_DAYS, min_periods=10).max()
    rolling_trough = close_ema.rolling(PEAK_TROUGH_LOOKBACK_DAYS, min_periods=10).min()

    dist_peak = (close_ema - rolling_peak) / rolling_peak * 100
    dist_trough = (close_ema - rolling_trough) / rolling_trough * 100

    events = []
    last_loss_tier = None
    last_buy_tier = None

    open_trade_entry_date = None
    open_trade_entry_price = None

    for i, dt in enumerate(df.index):
        if i < PEAK_TROUGH_LOOKBACK_DAYS:
            continue

        dp = dist_peak.iloc[i]
        dt_low = dist_trough.iloc[i]
        price = float(close_ema.iloc[i])

        # 손절 단계
        if dp <= ts["forced_pct"]:
            loss_tier = "3rd_forced"
        elif dp <= ts["alert_pct"]:
            loss_tier = "2nd_alert"
        elif dp <= ts["watch_pct"]:
            loss_tier = "1st_watch"
        else:
            loss_tier = None

        # 매수 단계
        if dt_low >= tb["add_pct"]:
            buy_tier = "3rd_add"
        elif dt_low >= tb["alert_pct"]:
            buy_tier = "2nd_alert"
        elif dt_low >= tb["watch_pct"]:
            buy_tier = "1st_watch"
        else:
            buy_tier = None

        # 단계가 새로 도달했을 때만 event 발생 (escalation only)
        loss_levels = ["1st_watch", "2nd_alert", "3rd_forced"]
        buy_levels = ["1st_watch", "2nd_alert", "3rd_add"]

        last_loss_idx = loss_levels.index(last_loss_tier) if last_loss_tier in loss_levels else -1
        cur_loss_idx = loss_levels.index(loss_tier) if loss_tier in loss_levels else -1
        last_buy_idx = buy_levels.index(last_buy_tier) if last_buy_tier in buy_levels else -1
        cur_buy_idx = buy_levels.index(buy_tier) if buy_tier in buy_levels else -1

        if cur_loss_idx > last_loss_idx and cur_loss_idx >= 1:  # 2차 이상
            entry_pnl = None
            if open_trade_entry_price is not None and loss_tier in ("2nd_alert", "3rd_forced"):
                entry_pnl = (price - open_trade_entry_price) / open_trade_entry_price * 100
                exit_event = {
                    "date": dt, "ticker": ticker, "type": "EXIT",
                    "tier": loss_tier, "price": price,
                    "ref_price": float(rolling_peak.iloc[i]),
                    "dist_pct": float(dp),
                    "entry_date": open_trade_entry_date,
                    "entry_price": open_trade_entry_price,
                    "trade_pnl_pct": entry_pnl,
                }
                events.append(exit_event)
                open_trade_entry_date = None
                open_trade_entry_price = None
            else:
                events.append({
                    "date": dt, "ticker": ticker, "type": "STOP_SIGNAL",
                    "tier": loss_tier, "price": price,
                    "ref_price": float(rolling_peak.iloc[i]),
                    "dist_pct": float(dp),
                    "entry_date": None, "entry_price": None,
                    "trade_pnl_pct": None,
                })

        if cur_buy_idx > last_buy_idx and cur_buy_idx >= 1:  # 2차 이상
            if open_trade_entry_price is None and buy_tier in ("2nd_alert", "3rd_add"):
                open_trade_entry_date = dt
                open_trade_entry_price = price
                events.append({
                    "date": dt, "ticker": ticker, "type": "ENTRY",
                    "tier": buy_tier, "price": price,
                    "ref_price": float(rolling_trough.iloc[i]),
                    "dist_pct": float(dt_low),
                    "entry_date": None, "entry_price": None,
                    "trade_pnl_pct": None,
                })
            else:
                events.append({
                    "date": dt, "ticker": ticker, "type": "BUY_SIGNAL",
                    "tier": buy_tier, "price": price,
                    "ref_price": float(rolling_trough.iloc[i]),
                    "dist_pct": float(dt_low),
                    "entry_date": None, "entry_price": None,
                    "trade_pnl_pct": None,
                })

        # tier 리셋: 가격이 임계값 밖으로 빠져나오면
        if dp > ts["watch_pct"]:
            last_loss_tier = None
        else:
            last_loss_tier = loss_tier
        if dt_low < tb["watch_pct"]:
            last_buy_tier = None
        else:
            last_buy_tier = buy_tier

    return pd.DataFrame(events) if events else pd.DataFrame()


def summarize(events_df: pd.DataFrame, ticker: str, df: pd.DataFrame) -> dict:
    """종목별 요약 통계."""
    if events_df.empty:
        return {"ticker": ticker, "n_events": 0}

    n_events = len(events_df)
    completed_trades = events_df[events_df["trade_pnl_pct"].notna()]
    n_trades = len(completed_trades)

    if n_trades > 0:
        avg_pnl = completed_trades["trade_pnl_pct"].mean()
        win_rate = (completed_trades["trade_pnl_pct"] > 0).mean() * 100
        max_gain = completed_trades["trade_pnl_pct"].max()
        max_loss = completed_trades["trade_pnl_pct"].min()
    else:
        avg_pnl = win_rate = max_gain = max_loss = np.nan

    # Buy & Hold 비교
    bh_return = (df['Close'].iloc[-1] / df['Close'].iloc[PEAK_TROUGH_LOOKBACK_DAYS] - 1) * 100

    return {
        "ticker": ticker,
        "n_events": n_events,
        "n_completed_trades": n_trades,
        "avg_trade_pnl_%": float(avg_pnl) if not pd.isna(avg_pnl) else None,
        "win_rate_%": float(win_rate) if not pd.isna(win_rate) else None,
        "max_gain_%": float(max_gain) if not pd.isna(max_gain) else None,
        "max_loss_%": float(max_loss) if not pd.isna(max_loss) else None,
        "buy_and_hold_%": float(bh_return),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--period", default="2y", help="yfinance period (e.g. 1y, 2y, 5y)")
    parser.add_argument("--ticker", default=None, help="단일 종목만 (없으면 전 종목)")
    parser.add_argument("--output", default="backtest_results.xlsx")
    args = parser.parse_args()

    tickers = [args.ticker] if args.ticker else list(TICKER_CONFIG.keys())

    all_events = []
    summaries = []
    price_data = {}

    for t in tickers:
        print(f"Backtesting {t}...")
        try:
            df = fetch_history(t, args.period)
            if df.empty:
                print(f"  no data, skipping")
                continue
            price_data[t] = df
            events = simulate(df, t)
            if not events.empty:
                all_events.append(events)
            summary = summarize(events, t, df)
            summaries.append(summary)
            print(f"  events: {summary['n_events']}, "
                  f"trades: {summary.get('n_completed_trades', 0)}, "
                  f"avg pnl: {summary.get('avg_trade_pnl_%')}")
        except Exception as e:
            print(f"  error: {e}")

    write_excel(all_events, summaries, price_data, args.output)
    print(f"\nResults written to {args.output}")


def write_excel(all_events, summaries, price_data, output_path):
    """결과를 Excel로 (Summary + Events + per-ticker tabs)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Summary
    ws = wb.create_sheet("Summary")
    headers = ["Ticker", "Events", "Trades", "Avg PnL %", "Win Rate %",
               "Max Gain %", "Max Loss %", "Buy&Hold %", "Strategy vs B&H"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Arial")
        c.fill = PatternFill("solid", start_color="2C3E50")
        c.alignment = Alignment(horizontal="center")

    for i, s in enumerate(summaries, 2):
        ws.cell(row=i, column=1, value=s["ticker"]).font = Font(bold=True, name="Arial")
        ws.cell(row=i, column=2, value=s.get("n_events", 0))
        ws.cell(row=i, column=3, value=s.get("n_completed_trades", 0))
        ws.cell(row=i, column=4, value=s.get("avg_trade_pnl_%"))
        ws.cell(row=i, column=5, value=s.get("win_rate_%"))
        ws.cell(row=i, column=6, value=s.get("max_gain_%"))
        ws.cell(row=i, column=7, value=s.get("max_loss_%"))
        ws.cell(row=i, column=8, value=s.get("buy_and_hold_%"))
        avg_pnl = s.get("avg_trade_pnl_%")
        bh = s.get("buy_and_hold_%")
        n_trades = s.get("n_completed_trades", 0)
        if avg_pnl is not None and bh is not None and n_trades:
            total_strat = avg_pnl * n_trades
            ws.cell(row=i, column=9, value=total_strat - bh)

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 10

    # Number format
    for row in range(2, len(summaries) + 2):
        for col in [4, 5, 6, 7, 8, 9]:
            ws.cell(row=row, column=col).number_format = '0.00'

    ws.cell(row=len(summaries) + 4, column=1,
            value="해석: 'Strategy vs B&H'가 양수면 strategy가 buy&hold보다 우수.")
    ws.cell(row=len(summaries) + 4, column=1).font = Font(italic=True, color="666666", name="Arial")

    # Sheet 2: All Events
    ws = wb.create_sheet("Events")
    if all_events:
        events_df = pd.concat(all_events, ignore_index=True)
        events_df = events_df.sort_values("date")
        cols = ["date", "ticker", "type", "tier", "price", "ref_price",
                "dist_pct", "entry_date", "entry_price", "trade_pnl_pct"]
        for col, h in enumerate(cols, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF", name="Arial")
            c.fill = PatternFill("solid", start_color="34495E")
        for r, (_, row) in enumerate(events_df.iterrows(), 2):
            for c, col in enumerate(cols, 1):
                val = row.get(col)
                if pd.isna(val):
                    val = None
                if col in ("date", "entry_date") and val is not None:
                    val = pd.Timestamp(val).strftime("%Y-%m-%d")
                ws.cell(row=r, column=c, value=val)
                if col in ("price", "ref_price", "entry_price"):
                    ws.cell(row=r, column=c).number_format = '$#,##0.00'
                elif col in ("dist_pct", "trade_pnl_pct"):
                    ws.cell(row=r, column=c).number_format = '0.00'

        for col in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

    # Per-ticker price + events tab with chart
    for ticker, df in price_data.items():
        ws = wb.create_sheet(f"{ticker}_chart")
        ws.cell(row=1, column=1, value="Date").font = Font(bold=True, name="Arial")
        ws.cell(row=1, column=2, value="Close").font = Font(bold=True, name="Arial")
        ws.cell(row=1, column=3, value="EMA5").font = Font(bold=True, name="Arial")

        close_ema = df['Close'].ewm(span=EMA_SMOOTHING_DAYS, adjust=False).mean()

        for i, (dt, row) in enumerate(df.iterrows(), 2):
            ws.cell(row=i, column=1, value=dt.strftime("%Y-%m-%d"))
            ws.cell(row=i, column=2, value=float(row['Close'])).number_format = '$#,##0.00'
            ws.cell(row=i, column=3, value=float(close_ema.iloc[i-2])).number_format = '$#,##0.00'

        chart = LineChart()
        chart.title = f"{ticker} — Price + 5d EMA"
        chart.y_axis.title = "Price ($)"
        chart.x_axis.title = "Date"
        chart.height = 10
        chart.width = 22

        data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=len(df) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E2")

        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 14

    wb.save(output_path)


if __name__ == "__main__":
    main()
